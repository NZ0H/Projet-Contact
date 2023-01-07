import openpyxl
import argparse
import os

####################################################
#    BIEN CE PLACER DANS LE REPERTOIRE TEST POUR   #
#       LA FONCTIONNALITE DU PROGRAMME             #
####################################################

valeur_tmp_fct2=[]

def argument():
    parser = argparse.ArgumentParser()
    parser.add_argument('--file',type=str,nargs='+')
    parser.add_argument('--sortie',type=str,nargs='+')
    parser.add_argument('--nom_page',type=str,nargs='+')
    args=parser.parse_args()
    return args

def fichier_excel(args):
    workbook = openpyxl.load_workbook('../data/Stages/'+ args, data_only = True)
    titres_onglets = workbook.sheetnames
    onglet = workbook[titres_onglets[0]]
    nb_none=0
    listes_data_depart = [] #stocke toute les lignes
    listes_data_traiter = []

    """data recupérer directement dans le excel, sans avoir à faire de trie"""
    data_nom_entreprise=[]
    data_lieu=[]
    data_sujet=[]
    data_tuteur=[]
    data_tel=[]
    data_mail=[]

    """ data ayant était traité"""
    data_code_postal=[]
    data_prenom_tuteur=[]
    data_nom_tuteur=[]
    data_civilite_tuteur=[]
    data_code_postal=[]
    data_ville=[]
    nom_prenom_tmp=[]
    nom_tmp=[]
    prenom_tmp=[]
    valeur_tmp=[]
    valeur=[]


    for row in onglet.values: 
        listes_data_depart.append(list(row))

    for r in range(len(listes_data_depart)):    
        for non in listes_data_depart[r]:       
            if non == None:
                nb_none+=1
        if len(listes_data_depart[r])==nb_none:
            listes_data_depart[r].clear()
        nb_none=0

    for valeur in listes_data_depart :
        if len(valeur)!=0:
            listes_data_traiter.append(valeur)

    for titre in range(len(listes_data_traiter[0])):
        if listes_data_traiter[0][titre]=="Organisme d'accueil":
            for colonne in range(len(listes_data_traiter)):
                data_nom_entreprise.append(listes_data_traiter[colonne][titre])
            data_nom_entreprise.remove("Organisme d'accueil")
        if listes_data_traiter[0][titre]=='Lieu':
            for colonne in range(len(listes_data_traiter)):
                data_lieu.append(listes_data_traiter[colonne][titre])
            data_lieu.remove('Lieu')
        if listes_data_traiter[0][titre]=='Sujet':
            for colonne in range(len(listes_data_traiter)):
                data_sujet.append(listes_data_traiter[colonne][titre])
            data_sujet.remove('Sujet')
        if listes_data_traiter[0][titre]=='Tuteur entreprise':
            for colonne in range(len(listes_data_traiter)):
                data_tuteur.append(listes_data_traiter[colonne][titre])
            data_tuteur.remove('Tuteur entreprise')
        if listes_data_traiter[0][titre]=='Contact tel':
            for colonne in range(len(listes_data_traiter)):
                data_tel.append(listes_data_traiter[colonne][titre])
            data_tel.remove('Contact tel')
        if listes_data_traiter[0][titre]=='Mail':
            for colonne in range(len(listes_data_traiter)):
                data_mail.append(listes_data_traiter[colonne][titre])
            data_mail.remove('Mail')

    workbook.close()

    for i in range(len(data_tuteur)):
        if data_tuteur[i]==None:
            nom_prenom_tmp.append(data_tuteur[i])
        else:
            nom_prenom_tmp.append(data_tuteur[i].split())

    for info in range(len(nom_prenom_tmp)):
            data_civilite_tuteur.append(nom_prenom_tmp[info][0])
            del nom_prenom_tmp[info][0]
            for n_p in range(len(nom_prenom_tmp[info])):
                if nom_prenom_tmp[info][n_p].isupper() == True :
                    nom_tmp.append(nom_prenom_tmp[info][n_p])
                else :
                    prenom_tmp.append(nom_prenom_tmp[info][n_p])

            data_nom_tuteur.append(" ".join(nom_tmp))
            nom_tmp.clear()
            data_prenom_tuteur.append(" ".join(prenom_tmp))
            prenom_tmp.clear()
    

    """separation code postal et ville"""
    for i in range(len(data_lieu)):
        data_code_postal.append('Champs vide')
        valeur_tmp.append(data_lieu[i].split())
        for v in range(len(valeur_tmp[i])):
            """cree liste avec valeur vide"""
            if valeur_tmp[i][v].isdecimal() == True :
                data_code_postal[i]=valeur_tmp[i][v]
            if valeur_tmp[i][v].isalpha() == True :
                valeur.append(valeur_tmp[i][v])
        data_ville.append(" ".join(valeur))
        valeur.clear()
                
    for i in range(len(data_nom_entreprise)):        
        if  len(data_civilite_tuteur)==0 or len(data_code_postal)==0 or len(data_mail)==0 or len(data_sujet)==0 or len(data_ville)==0 or len(data_tel)==0:
            

            if len(data_civilite_tuteur)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_civilite_tuteur.append('Non_renseigné')

            if len(data_code_postal)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_code_postal.append('Non_renseigné')

            if len(data_mail)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_mail.append('Non_renseigné')

            if len(data_sujet)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_sujet.append('Non_renseigné')

            if len(data_ville)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_ville.append('Non_renseigné')

            if len(data_tel)==0 :
                for rep in range(len(data_nom_entreprise)):
                    data_tel.append('Non_renseigné')

    print(len(data_nom_entreprise),len(data_ville),len(data_code_postal),len(data_sujet),len(data_civilite_tuteur),len(data_nom_tuteur),len(data_prenom_tuteur),len(data_tel),len(data_mail))

    return data_nom_entreprise,data_ville,data_code_postal,data_sujet,data_civilite_tuteur,data_nom_tuteur,data_prenom_tuteur,data_tel,data_mail


def contact(args_sortie):
    all_value=valeur_tmp_fct2

    data_nom_entreprise=[]
    data_ville=[]
    data_code_postal=[]
    data_sujet=[]
    data_civilite_tuteur=[]
    data_nom_tuteur=[]
    data_prenom_tuteur=[]
    data_tel=[]
    data_mail=[]

    for fichier in range(len(all_value)):
        for liste in range(len(all_value[fichier])):
            if liste == 0 :
                for valeur in all_value[fichier][liste]:
                    data_nom_entreprise.append(valeur)
            if liste == 1 :
                    for valeur in all_value[fichier][liste]:
                        data_ville.append(valeur)
            if liste == 2 :
                    for valeur in all_value[fichier][liste]:
                        data_code_postal.append(valeur)
            if liste == 3 :
                    for valeur in all_value[fichier][liste]:
                        data_sujet.append(valeur)
            if liste == 4 :
                    for valeur in all_value[fichier][liste]:
                        data_civilite_tuteur.append(valeur)
            if liste == 5 :
                    for valeur in all_value[fichier][liste]:
                        data_nom_tuteur.append(valeur)

            if liste == 6 :
                for valeur in all_value[fichier][liste]:
                    data_prenom_tuteur.append(valeur)
            if liste == 7 :
                for valeur in all_value[fichier][liste]:
                    data_tel.append(valeur)
            if liste == 8 :
                for valeur in all_value[fichier][liste]:
                    data_mail.append(valeur)
    print(len(data_nom_entreprise))
    print(len(data_nom_entreprise),len(data_ville),len(data_code_postal),len(data_sujet),len(data_civilite_tuteur),len(data_nom_tuteur),len(data_prenom_tuteur),len(data_tel),len(data_mail))

    if args_sortie[len(args_sortie)-4:] == 'xlsx':
        wb_out = openpyxl.Workbook()
        ws1 = wb_out.active
        ws1.title = "Etudiants"
        ws1.append(["Nom Entreprise","Ville","Code Postal","Sujet","Civilité","Nom","Prénom","Téléphone","Email"])
        for colonne1,colonne2,colonne3,colonne4,colonne5,colonne6,colonne7,colonne8,colonne9 in zip(data_nom_entreprise,data_ville,data_code_postal,data_sujet,data_civilite_tuteur,data_nom_tuteur,data_prenom_tuteur,data_tel,data_mail):
            ws1.append([colonne1,colonne2,colonne3,colonne4,colonne5,colonne6,colonne7,colonne8,colonne9])

        wb_out.save('../contact/'+args_sortie)

    if args_sortie[len(args_sortie)-5:] == 'vcard':
        
        for num_c in range(len(data_nom_tuteur)):
           
            
                

            print(len(data_nom_entreprise),len(data_ville),len(data_code_postal),len(data_sujet),len(data_civilite_tuteur),len(data_nom_tuteur),len(data_prenom_tuteur),len(data_tel),len(data_mail))
            f= open("../contact/Dossier_VCARD/vcard_{}_{}.vcf".format(data_nom_tuteur[num_c],data_prenom_tuteur[num_c]),"w")

            f.write("BEGIN:VCARD\nVERSION:3.0\n")
            f.write("FN:{} {}\n".format(data_prenom_tuteur[num_c],data_nom_tuteur[num_c]))
            f.write("N:{};{};;;\n".format(data_nom_tuteur[num_c],data_prenom_tuteur[num_c]))
            f.write("item1.EMAIL;TYPE=INTERNET:{}\n".format(data_mail[num_c]))
            f.write("item1.X-ABLabel:\nitem2.TEL:{}\n".format(data_tel[num_c]))
            f.write("item2.X-ABLabel:\nitem3.ORG:{}\n".format(data_nom_entreprise[num_c]))
            f.write("item3.X-ABLabel:\nNOTE:{}\n".format(data_sujet[num_c]))
            f.write("CATEGORIES:myContacts\nEND:VCARD")
            print('Ca marche')

print(contact('fichier_contact.xlsx'))


def page_web(all_values):
    nom_fichier = "html"
    os.makedirs(nom_fichier, exist_ok=True)

    structure = ["Entreprise", "Civilite", "Nom", "Prenom", "Email", "Telephone"]
    
    # Extract the data from the xlsx file
    data = all_values
    # Créer la structure de la page web
    f = open(nom_fichier+"/html.html","w")
    f.write("<!DOCTYPE html>\n")
    f.write("<head> \n<meta charset='utf-8'>\n")
    f.write("<link rel='stylesheet' href='style.css'>")
    f.write("\n<title> Tableau </title> \n</head>")
    f.write("\n<body>\n")

    # Créer le tableau
    f.write("<table>\n<tr>")

    for i in structure:
        f.write("<th>{}</th>".format(i))
    f.write("</tr>") 

    # Ecris tout les donnée dans le tableau
    for row in data:
        f.write("<tr>")
        for cell in row:
            f.write("<td>{}</td>".format(cell))
        f.write("</tr>")

    f.write("</table>\n</body>\n</html>")
    f.close()


arguments=argument()

for nb_arg in arguments.file :
    valeur_tmp_fct2.append(fichier_excel(nb_arg))
    print('FInit')

for nb_arg in arguments.sortie :
    print(nb_arg)
    contact(nb_arg)
    print('test réussi')
