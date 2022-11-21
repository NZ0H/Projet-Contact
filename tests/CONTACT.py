import selenium #https://selenium-python.readthedocs.io/
import openpyxl as xl #https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook


#./nom_projet.py --input-file fichier_annee1.xls fichier_annee2.xls  --contacts liste_contacts.vcf --output-dir ./../html

def input(file):
    """
    Selctionner les fichiers voulant etre traité

    """
    #importation du fichier à traiter
    wb=xl.load_workbook('stage 2005.xlsx')

    #nom des feuilles contenue dans le fichier
    nom_feuille=wb.sheetnames

    #affiche la taille
    for i in nom_feuille:
        sheet=wb[i]
        print(sheet.calculate_dimension())

    pass

def contact():
    """
    Convertir la liste des fichiers séléctionné et en faire un seul fichier au format vcf
    """
    pass

def output(dir):
    """"
    Création du tableau dans le fichier html
    """
    pass

#affiche la taille d'un tableau sheet.calculate_dimension()

#importation du fichier à traiter
wb=xl.load_workbook('stage 2005.xlsx')

#nom des feuilles contenue dans le fichier
nom_feuille=wb.sheetnames


liste_titre_tableau=[]
#test1
"""
for i in nom_feuille:
    sheet=wb[i]
    for row in sheet.values:
        for titre in row :
            if titre=="NOM " or titre=="Prénom":
                print(titre)
"""

for i in nom_feuille:
    sheet=wb[i]
    for row in (sheet.columns,sheet.values):
        for a in row :
            print(a)

"""
    for j, col in enumerate(sheet.columns, 1):
        print(values)


    for row in sheet.values:
        for titre in row :
            if titre=="NOM " or titre=="Prénom":
                print(sheet.cell(t))

"""









