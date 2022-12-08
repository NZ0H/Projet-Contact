import openpyxl
import vcf

#Fichier pour tester la suppression des lignes None

def input(file):
    workbook = openpyxl.load_workbook(file, data_only = True)
    titres_onglets = workbook.sheetnames
    onglet = workbook[titres_onglets[0]]

    nb_none=0

    listes_data_depart = [] #stocke toute les lignes
    listes_data_traiter = []

    data_nom_entreprise=[]
    data_lieu=[]
    data_sujet=[]
    data_tuteur=[]
    data_tel=[]
    data_mail=[]
    data_code_postal=[]


    for row in onglet.values: #ajoute les lignes dans la liste lignes

        listes_data_depart.append(list(row))

    for r in range(len(listes_data_depart)):        #parcourt les listes présente dans lignes

        for non in listes_data_depart[r]:       #parcourt liste de liste pour avoir les valeurs individuelles
            if non == None:
                nb_none+=1


        if len(listes_data_depart[r])==nb_none:     # vide la liste remplie de non
            listes_data_depart[r].clear()
        nb_none=0

    for valeur in listes_data_depart :
        if len(valeur)!=0:
            listes_data_traiter.append(valeur)

    for titre in range(len(listes_data_traiter[0])):

        if listes_data_traiter[0][titre]=="Organisme d'accueil":
            for colonne in range(len(listes_data_traiter)):
                data_nom_entreprise.append(listes_data_traiter[colonne][titre])

        if listes_data_traiter[0][titre]=='Lieu':  # Si le nom de la colonne égal ADRENTR1 alors on parcourt ligne ADRENTR1 et on ajoute chaque adrsse d'entreprise dans la liste data_adresse
            for colonne in range(len(listes_data_traiter)):
                print(titre, listes_data_traiter[colonne][titre])
                data_lieu.append(listes_data_traiter[colonne][titre])

        if listes_data_traiter[0][titre]=='Sujet':                          # Si le nom de la colonne égal LOCALITE alors on parcourt ligne LOCALITE et on ajoute chaque ville dans la liste data_ville
            for colonne in range(len(listes_data_traiter)):
                data_sujet.append(listes_data_traiter[colonne][titre])

        if listes_data_traiter[0][titre]=='Tuteur entreprise':                             # Si le nom de la colonne égal Pays alors on parcourt ligne Pays et on ajoute chaque pays dans la liste data_pays
            for colonne in range(len(listes_data_traiter)):
                data_tuteur.append(listes_data_traiter[colonne][titre])

        if listes_data_traiter[0][titre]=='Contact tel':
            for colonne in range(len(listes_data_traiter)):
                data_tel.append(listes_data_traiter[colonne][titre])

        if listes_data_traiter[0][titre]=='Mail':
            for colonne in range(len(listes_data_traiter)):
                data_mail.append(listes_data_traiter[colonne][titre])


    workbook.close()
    return data_nom_entreprise,data_lieu,data_sujet,data_tuteur,data_tel,data_mail

valeur_tmp=input('Entreprises_stage_2020.xlsx')



def contact(new_file):
    """
    Convertir la liste des fichiers séléctionné et en faire un seul fichier au format vcf

    Toute les colones à faire dans le fichiers .xlsx
    colone 2 : nom_entreprise
    colone 3 : adresse
    colone 4 : ville
    colone 5 : code_postal

    Toute les colones à faire dans le fichiers .xlsx
    colone nom et prenom
    colonne entreprise
    colonne fonction
    colonne email
    colonne e-mail
    colonne téléphone



data_nom_entreprise=[]
    data_lieu=[]
    data_sujet=[]
    data_tuteur=[]
    data_tel=[]
    data_mail=[]
    data_code_postal=[]


    """
    all_value = valeur_tmp

    wb_out = openpyxl.Workbook()
    #accès à la première feuille du fichier
    ws1 = wb_out.active


# ajout entête des colonnes
    ws1.append(["Nom", "Prénom", "Note"])
    donnees = [
   ['Nom1', 'Prénom1', 14],
   ['Nom2', 'Prénom2', 12],
   ['Nom3', 'Prénom3', 11],
   ['Nom4', 'Prénom4', 17],
   ['Nom5', 'Prénom5', 8]]

    for row in all_value:
        ws1.append(row)

    #écriture du fichier
    wb_out.save(filename = new_file)

    return all_value

print(contact('test.xlsx'))



vf=vcf.VCFWriter()
"""

    workbook = openpyxl.load_workbook(file, data_only = True)

    titres_onglets = workbook.sheetnames
    onglet1 = workbook[titres_onglets[0]]
    print(onglet1.calculate_dimension())
    nb_none=0
    nb_liste_none=0
    data_principal=[]



    for row in onglet1.iter_rows(min_row=0,max_row=30, values_only=True):
        data_principal.append(row)  #regroupe toute les valeurs dans le tableau data_principal


    for info in range(len(data_principal)) :    #parcours des listes dans la liste data_principal

        for non in range(len(data_principal[info])):


            if data_principal[info][non] == None:
                nb_none+=1


        if len(data_principal[info])==nb_none:     # si len = nb_none alors la liste est supprimé
            data_principal.pop(info)
            nb_liste_none+=1

        nb_none=0
    """










